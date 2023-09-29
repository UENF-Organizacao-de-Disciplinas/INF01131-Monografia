import os
import openpyxl
from copy import copy

# Obtém o caminho do diretório atual
caminho_pasta = os.getcwd()

# Lista todos os arquivos na pasta atual
arquivos = os.listdir(caminho_pasta)

# Filtra apenas os arquivos Excel
arquivos_excel = [arquivo for arquivo in arquivos if arquivo.endswith('.xlsx')]

# Cria uma nova pasta de trabalho para armazenar todas as guias
wb_novo = openpyxl.Workbook()

# Loop através dos arquivos
for arquivo in arquivos_excel:
    # Caminho completo do arquivo
    caminho_arquivo = os.path.join(caminho_pasta, arquivo)

    # Carrega o arquivo atual
    wb_atual = openpyxl.load_workbook(caminho_arquivo)

    # Loop através das guias do arquivo atual
    for nome_guia in wb_atual.sheetnames:
        # Seleciona a guia atual
        ws_atual = wb_atual[nome_guia]

        # Cria uma nova guia na nova pasta de trabalho com o mesmo nome da guia atual
        ws_novo = wb_novo.create_sheet(title=nome_guia)

        # Copia os dados e a formatação da guia atual para a nova guia
        for row in ws_atual:
            for cell in row:
                ws_novo[cell.coordinate].value = cell.value
                if cell.has_style:
                    ws_novo[cell.coordinate]._style = copy(cell._style)

# Salva a nova pasta de trabalho
wb_novo.save('novo_arquivo.xlsx')
